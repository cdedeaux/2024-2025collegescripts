from openpyxl import *
import pandas as pd
wb = Workbook()
ws = wb.active
wb = load_workbook("Non-Greek.xlsx")
ws = wb.active

emails = []

for cell in ws["B"][1:]:
    emails.append(str(cell.value) + "@msstate.edu")

df = pd.DataFrame(emails, columns=["Email"])
df.to_csv("Non-Greek-Emails.csv", index=False)